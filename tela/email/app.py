## python -m venv venv
##  virtualvenv venv
## venv/Scripts/Activate.ps1
## pip install pysimplegui
## pip install openpyxl


import PySimpleGUI as sg
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import datetime as dt 
from openpyxl import Workbook, load_workbook


hoje = dt.date.today()

# recuperando dados da conta
from func import principal
host            = principal()[0]
port            = principal()[1]
password        = principal()[2] 
remetente       = principal()[3] 
user            = principal()[4] 

# Configurando tema padrão
sg.theme('SystemDefault1')  


# Dados de caminho e arquivo contendo lista para disparao de e-mails
caminho = "C:\\py\\dados\\fin\\cobranca\\"
arquivo = "lista.xlsx"



# função para disparar e-mail e atualizar excel
def autaliza():
    wb = load_workbook(f'{caminho}{arquivo}') 
    sheet_obj = wb.active
    cont = 2

    

    for i in sheet_obj:
        if sheet_obj.cell(row = cont, column = 1).value != None:
            indice = sheet_obj.cell(row = cont, column = 1).value
            pessoa = sheet_obj.cell(row = cont, column = 2).value
            dest = sheet_obj.cell(row = cont, column = 3).value
            message = sheet_obj.cell(row = cont, column = 4).value 
            enviado = sheet_obj.cell(row = cont, column = 5).value
            sheet_obj.cell(row = cont, column = 5).value = hoje
        
            cont= cont + 1

            if enviado == None:
                server      = smtplib.SMTP(host, port)
                server.ehlo()
                server.starttls()
                server.login(user, password)
                email_msg   = MIMEMultipart()
                email_msg['From'] = remetente
                email_msg['To'] = dest
                email_msg['Subject'] = 'Email testes by Python'
                email_msg.attach(MIMEText(message, 'plain'))
                server.sendmail(email_msg['From'], email_msg['To'], email_msg.as_string())
                server.quit()
    wb.save(f'{caminho}{arquivo}')
    
    listar()



# função para carregar dados para a tabela
def listar():
    registros = []
    wb = load_workbook(f'{caminho}{arquivo}') 
    sheet_obj = wb.active
    cont = 2
   
    for i in sheet_obj:
        if sheet_obj.cell(row = cont, column = 1).value != None:
            indice = sheet_obj.cell(row = cont, column = 1).value
            pessoa = sheet_obj.cell(row = cont, column = 2).value
            email = sheet_obj.cell(row = cont, column = 3).value
            if sheet_obj.cell(row = cont, column = 5).value == None:
                enviado = ''
            else:
                enviado = sheet_obj.cell(row = cont, column = 5).value
            
            cont= cont + 1
            result = [indice, pessoa, email,enviado]
            registros.append(result)
    
    return registros




# função disfarce para indicar que o processamento foi incluido - ela só existe pois não consegui atualizar a tabela com refrash
def tela_2():
    sg.theme('SystemDefault1')
    layout = [
        [sg.Text("E-mails enviados", font="Arial 16 bold", pad=(25, 30))],
        [sg.Button("Finalizar", key="consulta", size=(40, 2))]
    ]
    janela = sg.Window("Hermes Kairós", layout, text_justification="center")
    botao, valores = janela.read()
    janela.close()
    return botao





# Carregando dados para grade
def t(dados):
    sg.set_options(font=("Arial Bold", 14))
    toprow = ['#', 'Pessoa', 'E-mail','Enviado']
    rows = [[dados]]

    # Definindo grade
    tbl1 = sg.Table(values=dados, headings=toprow,
    auto_size_columns=True,
    display_row_numbers=False,
    justification='center', key='-TABLE-',
    selected_row_colors='red on yellow',
    enable_events=True,
    expand_x=True,
    expand_y=True,
    enable_click_events=True)
    return tbl1




# Layout da tela da aplicação
layout = [  [sg.Text('Hermes é o Deus menssageiro e Kairós o Deus das oportunidades!')],
            [sg.Text('Junstos entregam as mensagem pela Newton!')],
            [sg.Text('')],
            [sg.Text('Remetente para retorno: '), sg.Input(key='remet')],
            [sg.Text('')],
            [[t(listar())]],
            [sg.Button('Enviar'), sg.Button('Cancel')] ]

# Creando jenela Window
window = sg.Window('Hermes Kairós', layout,  size=(915, 500), resizable=True)
# Laço infinito
data1 = [[hoje,'9','8']]
while True:
    event, values = window.read()
    if event == sg.WIN_CLOSED or event == 'Cancel': 
        break
    if event == sg.WIN_CLOSED or event == 'Enviar':
        autaliza()
        window.close()
        tela_2()
  
        

window.close()
